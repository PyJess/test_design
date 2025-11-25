# import pandas as pd
# import os
# from pathlib import Path

# def fix_labels_with_order(file_input):
    
#     file_input = Path(file_input)
#     output = file_input.with_name(file_input.stem + "_excel_fixed.xlsx")


#     mapping = { 
#         "Title":"Title",    
#         "#":"#",    
#         "Step":"Step",
#         "Step Description":"Step Description",
#         "Step Expected Result":"Expected Result",
#         "Test Group": "Test Group",
#         "Test Stage":"Test Stage",  
#         "Channel": "Canale",
#         "Device": "Dispositivo",
#         "Reference System": "Sistema Integrale di riferimento",
#         "Preconditions": "Precondizioni",
#         "Execution Mode": "Modalità Operativa",
#         "Functionality": "Funzionalità",
#         "Test Type": "Tipologia Test",
#         "No Regression Test": "Test di no regression",
#         "Partial Coverage Description": "Description Partial Coverage",
#         "Dataset":"Dataset",
#         "Expected Result":"Risultato Atteso",
#         "_polarion":"Linked Work Items",
#         "Priority":"Priority"
#     }


#     col_order = [
#         "Title",
#         "#",
#         "Step",
#         "Step Description",
#         "Expected Result",
#         "Test Group",
#         "Canale",
#         "Dispositivo",
#         "Priority",
#         "Test Stage",
#         "Sistema Integrale di riferimento",
#         "Precondizioni",
#         "Modalità Operativa",
#         "Funzionalità",
#         "Tipologia Test",
#         "Test di no regression",
#         "Automation",
#         "Dataset",
#         "Risultato Atteso",
#         "Linked Work Items"
#     ]

#     # Carica file
#     df = pd.read_excel(file_input)

#     # Applica mapping
#     df = df.rename(columns=mapping)

#     # Aggiunge eventuali colonne mancanti
#     for col in col_order:
#         if col not in df.columns:
#             df[col] = None

#     # Riordina
#     df = df[col_order]

#     # Salva
#     df.to_excel(output, index=False)

#     return df


# # test = Path(__file__).parent /"test_precondizioni_Dario_test_feedbackAI_precondizioni.xlsx"

# # fix_labels_with_order(test)

from openpyxl import load_workbook
from pathlib import Path

def fix_labels_with_order(file_input):

    file_input = Path(file_input)
    output = file_input.with_name(file_input.stem + "_excel_fixed_test_design.xlsx")

    # Mapping (vecchia colonna -> nuova colonna)
    mapping = { 
        "Title":"Title",    
        "#":"#",    
        "Step":"Step",
        "Step Description":"Step Description",
        "Step Expected Result":"Expected Result",
        "Test Group": "Test Group",
        "Test Stage":"Test Stage",  
        "Channel": "Canale",
        "Device": "Dispositivo",
        "Reference System": "Sistema Integrale di riferimento",
        "Preconditions": "Precondizioni",
        "Execution Mode": "Modalità Operativa",
        "Functionality": "Funzionalità",
        "Test Type": "Tipologia Test",
        "No Regression Test": "Test di no regression",
        "Partial Coverage Description": "Description Partial Coverage",
        "Dataset":"Dataset",
        "Expected Result":"Risultato Atteso",
        "_polarion":"Linked Work Items",
        "Priority":"Priority"
    }

    # Nuovo ordine colonne
    col_order = [
        "Title",
        "#",
        "Step",
        "Step Description",
        "Expected Result",
        "Test Group",
        "Canale",
        "Dispositivo",
        "Priority",
        "Test Stage",
        "Sistema Integrale di riferimento",
        "Precondizioni",
        "Modalità Operativa",
        "Funzionalità",
        "Tipologia Test",
        "Test di no regression",
        "Automation",
        "Dataset",
        "Risultato Atteso",
        "Linked Work Items"
    ]

    # --- 1. Apri il file con OpenPyXL mantenendo il formatting
    wb = load_workbook(file_input)
    ws = wb.active

    # --- 2. Leggi header attuale
    headers_old = [cell.value for cell in ws[1]]

    # --- 3. Crea un dizionario: nuova_colonna -> vecchia_colonna
    new_to_old = {}
    for new_col in col_order:
        for old_col in headers_old:
            if old_col in mapping and mapping[old_col] == new_col:
                new_to_old[new_col] = old_col
            elif old_col == new_col:
                new_to_old[new_col] = old_col

    # --- 4. Costruisci un nuovo foglio con colonne riordinate
    ws_new = wb.create_sheet("fixed")
    
    # Header nuovo
    for col_idx, col_name in enumerate(col_order, start=1):
        ws_new.cell(row=1, column=col_idx, value=col_name)

    # --- 5. Copia i dati e la formattazione cella → cella
    for row_idx in range(2, ws.max_row + 1):
        for col_idx, new_col in enumerate(col_order, start=1):

            if new_col not in new_to_old:
                continue

            old_col = new_to_old[new_col]
            old_col_index = headers_old.index(old_col) + 1

            old_cell = ws.cell(row=row_idx, column=old_col_index)
            new_cell = ws_new.cell(row=row_idx, column=col_idx)

            # copia valore
            new_cell.value = old_cell.value

            # copia formattazione (COLORI, FONT ecc.)
            if old_cell.has_style:
                new_cell._style = old_cell._style

    # --- 6. Rimuovi vecchio sheet e rinomina quello nuovo
    wb.remove(ws)
    ws_new.title = ws.title

    # --- 7. Salva il file FIXED
    wb.save(output)

    return output

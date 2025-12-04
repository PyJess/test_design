import json
import os
import pandas as pd
from collections import defaultdict
import subprocess
import re
from pathlib import Path
from typing import Dict, List, Tuple, Any
#from llm.llm import a_invoke_model
from langchain_openai import ChatOpenAI
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl import load_workbook
from openpyxl.cell.rich_text import TextBlock, CellRichText
from openpyxl.cell.text import InlineFont
from docx import Document
import base64
from docx.oxml import parse_xml

def load_file(filepath:str):
    with open(filepath, encoding="utf-8") as f:
        return f.read()


def load_json(filepath:str):
    with open(filepath, encoding="utf-8") as f:
        return json.load(f)


def save_json(data, path):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=4)


def get_user_path(user_id: str, subfolder: str = "") -> str:
    """Get the user-specific path."""
    base_dir = os.path.join(os.path.dirname(__file__),"outputs", user_id, subfolder)
    os.makedirs(base_dir, exist_ok=True)
    return base_dir


def excel_to_json(path):
    df = pd.read_excel(path)
    df = df.dropna(axis=1, how="all")  
    df = df.ffill()     
    df.columns = [str(c).strip() for c in df.columns]  

    step_cols = [c for c in df.columns if "step" in c.lower() or "result" in c.lower()]
    id_col = next((c for c in df.columns if c.lower() == "id"), None)

    if not id_col:
        raise ValueError("❌ Nessuna colonna 'ID' trovata nel file Excel!")

    # Tutte le altre colonne tranne gli step
    meta_cols = [c for c in df.columns if c not in step_cols]

    tests = {}

    for _, row in df.iterrows():
        test_id = str(row[id_col]).strip()
        if test_id not in tests:
            meta_data = {col: row.get(col, "") for col in meta_cols if col != id_col}
            meta_data["Steps"] = []
            tests[test_id] = meta_data

        step_data = {col: row.get(col, "") for col in step_cols}
        tests[test_id]["Steps"].append(step_data)

    output_dir = os.path.join(os.path.dirname(__file__),"..", "input")
    output_path = os.path.join(output_dir, "tests_output.json")

    # Scrive il file JSON
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(tests, f, indent=2, ensure_ascii=False)

    print(f"JSON creato in: {output_path}")
    return tests

PANDOC_EXE = "pandoc"
def process_docx(docx_path, output_base):
    """
    Process a DOCX file using Pandoc and split it into sections based on Markdown headers (#, ##, etc.).
    """
    print("processing pandoc")
    txt_output_path = os.path.join(output_base, Path(docx_path).stem + ".txt")
   
    docx_path = os.path.normpath(docx_path)
    txt_output_path = os.path.normpath(txt_output_path)
    os.makedirs(output_base, exist_ok=True)
   
    # Convert in md
    command = [
        PANDOC_EXE,
        "-s", docx_path,
        "--columns=120",
        "-t", "markdown",
        "-o", txt_output_path
    ]
   
    try:
        subprocess.run(command, check=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE, timeout=100)
    except subprocess.TimeoutExpired:
        print(f"[ERROR] Pandoc conversion timed out after 120 seconds")
        return [], []
    except Exception as e:
        print(f"[ERROR] Pandoc conversion failed: {e}")
        return [], [] 
    
    with open(txt_output_path, "r", encoding="utf-8") as f:
        text_lines = f.read().splitlines()
   
    headers = []
    heading_list = []
   
    for index, line in enumerate(text_lines):
        if line.startswith("#"):
            level = line.count("#")
            clean_name = line.replace("#", "").strip()
            headers.append([clean_name, index, level])
            heading_list.append([clean_name, level])
    
    
    headers.insert(0, ["== first line ==", 0, 0])
    headers.append(["== last line ==", len(text_lines), 0])
    heading_list.insert(0, ["== first line ==", 0])
    heading_list.append(["== last line ==", 0])
    
    head=[]
    chunks = []
    for i in range(len(headers) - 1):
        start_idx = headers[i][1]
        end_idx = headers[i + 1][1]
        #section_lines = text_lines[start_idx:end_idx]
        if headers[i][0] == "== first line ==":
            section_lines = text_lines[start_idx:end_idx]
        else:
            section_lines = text_lines[start_idx + 1:end_idx]
        chunk_text = "\n".join(section_lines).strip()
       
        header_cleaned = re.sub(r"\s*\{.*?\}", "", headers[i][0])
        header_cleaned = header_cleaned.replace("--", "–").strip(" *[]\n")
        #chunk_text = header_cleaned + "\n" + chunk_text
        chunk_text = chunk_text
        head.append(header_cleaned)
        chunks.append(chunk_text)
        #print(f" Paragrafo {i+1}: {chunk_text} \n\n")
        print(f" paragrafo {i+1} {header_cleaned}")
    
    return chunks, head


def process_docx_with_image(docx_path, output_base):
    """
    Process a DOCX file using Pandoc and split it into sections based on Markdown headers.
    Also extracts embedded images and maps them to their respective sections.
    """
    print("processing pandoc")
    
    images_folder = os.path.join(output_base, "Images", Path(docx_path).stem)
    os.makedirs(images_folder, exist_ok=True)
    txt_output_path = os.path.join(output_base, Path(docx_path).stem + ".md")
   
    docx_path = os.path.normpath(docx_path)
    txt_output_path = os.path.normpath(txt_output_path)
    images_folder = os.path.normpath(images_folder)
    os.makedirs(output_base, exist_ok=True)
   
    # Convert to markdown with image extraction
    command = [
        PANDOC_EXE,
        "-s", docx_path,
        "--columns=120",
        "-t", "markdown",
        "-o", txt_output_path,
        f"--extract-media={images_folder}"  # Aggiunto per estrarre le immagini
    ]
   
    try:
        subprocess.run(command, check=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE, timeout=30)
    except Exception as e:
        print(f"[ERROR] Pandoc conversion failed: {e}")
        return [], [], {}, []
    
    with open(txt_output_path, "r", encoding="utf-8") as f:
        text_lines = f.read().splitlines()
   
    headers = []
    heading_list = []
   
    for index, line in enumerate(text_lines):
        if line.startswith("#"):
            level = line.count("#")
            clean_name = line.replace("#", "").strip()
            headers.append([clean_name, index, level])
            heading_list.append([clean_name, level])
    
    headers.insert(0, ["== first line ==", 0, 0])
    headers.append(["== last line ==", len(text_lines), 0])
    heading_list.insert(0, ["== first line ==", 0])
    heading_list.append(["== last line ==", 0])
    
    head = []
    chunks = []
    for i in range(len(headers) - 1):
        start_idx = headers[i][1]
        end_idx = headers[i + 1][1]
        section_lines = text_lines[start_idx:end_idx]
        chunk_text = "\n".join(section_lines).strip()
       
        header_cleaned = re.sub(r"\s*\{.*?\}", "", headers[i][0])
        header_cleaned = header_cleaned.replace("--", "–").strip(" *[]\n")
        chunk_text = header_cleaned + "\n" + chunk_text
        head.append(header_cleaned)
        chunks.append(chunk_text)
        print(f" paragrafo {i+1} {header_cleaned}")
    
    # Gestione immagini
    images_dict = {}
    images_per_chunk = [[] for _ in chunks]
    available_images = {}
    
    # Pandoc crea una cartella "media" dentro images_folder
    media_folder = os.path.join(images_folder, "media")
    if os.path.exists(media_folder):
        images_folder = media_folder
        print(f"[Info] Using media folder: {images_folder}")
    
    # Cataloga le immagini disponibili
    if os.path.exists(images_folder):
        print(f"\n[DEBUG] Available images in {images_folder}:")
        for file in os.listdir(images_folder):
            print(f"  - {file}")
            if file.lower().endswith(('.png', '.jpg', '.jpeg', '.gif', '.bmp', '.webp', '.svg')):
                available_images[file.lower()] = file
    
    # Pattern per trovare riferimenti alle immagini nel markdown
    image_pattern = re.compile(
        r'!\[[^\]]*\]\([^)]*((?:img|image)\d+\.(?:png|jpg|jpeg|gif|bmp|webp|svg))[^)]*\)',
        re.IGNORECASE
    )
    
    # Mappa immagini ai chunk
    for i, chunk in enumerate(chunks):
        found = set()
        for match in image_pattern.finditer(chunk):
            img_filename = match.group(1).lower()
            if img_filename in found or img_filename not in available_images:
                continue
            
            actual_name = available_images[img_filename]
            img_path = os.path.join(images_folder, actual_name)
            
            # Encode immagine in base64
            try:
                with open(img_path, 'rb') as img_file:
                    encoded = base64.b64encode(img_file.read()).decode('utf-8')
                    images_dict[actual_name] = encoded
                    images_per_chunk[i].append(actual_name)
                    found.add(img_filename)
                    print(f"[Info] Chunk {i} → added image {actual_name}")
            except Exception as e:
                print(f"[Error] Failed to encode image {actual_name}: {e}")
    
    print(f"[Summary] Processed {len(chunks)} sections, found {len(images_dict)} unique images.")
    
    return chunks, head, images_dict, images_per_chunk


def apply_red_text(cell):
    """Color text in [[RED]]...[[/RED]] red, preserving the rest."""
    text = str(cell.value)
    if "[[RED]]" not in text:
        return  

    parts = re.split(r'(\[\[RED\]\]|\[\[/RED\]\])', text)
    rich_text = CellRichText()

    red = False
    for part in parts:
        if part == "[[RED]]":
            red = True
        elif part == "[[/RED]]":
            red = False
        elif part:
            font = InlineFont(color="FF0000") if red else InlineFont(color="000000")
            rich_text.append(TextBlock(font, part))

    cell.value = rich_text



def fill_excel_file(test_cases: dict):
    """
    Salva i test case in Excel, mantenendo gli step su righe separate
    e applica i testi rossi dove necessario.
    """
    field_mapping = {
        'Canale': 'Channel',
        'Dispositivo': 'Device',
        'Sistema di riferimento': 'Reference System',
        'Modalità Operativa': 'Execution Mode',
        'Funzionalità': 'Functionality',
        'Tipologia Test': 'Test Type',
        'Test di no regression': 'No Regression Test',
        'Automation': 'Automation',
        'Risultato Atteso': 'Expected Result',
        '_polarion': '_polarion'
    }
    
    # Definisci colonne finali (solo inglese)
    columns = [
        'Title', 'ID', '#', 'Test Group', 'Channel', 'Device', 
        'Priority', 'Test Stage', 'Reference System', 
        'Preconditions', 'Execution Mode', 'Functionality', 
        'Test Type', 'No Regression Test', 'Automation',
        'Dataset', 'Expected Result', 
        'Step', 'Step Description', 'Step Expected Result',
        'Country', 'Project', 'Author', 'Assignee(s)', 'Type', 
        'Partial Coverage Description', '_polarion',
        'Analysis', 'Coverage', 'Dev Complexity', 'Execution Time', 
        'Volatility', 'Developed', 'Note', 'Team Ownership', 
        'Team Ownership Note', 'Requires Script Maintenance'
    ]

    rows = []
    for tc_id, tc_data in test_cases.items():
        steps = tc_data.get('Steps', [])
        
        if not steps:
            steps = [{}]
        
        first = True
        for step in steps:
            row = {}

            if first:

                for col in columns:
                    if col not in ['Step', 'Step Description', 'Step Expected Result']:
                        value = tc_data.get(col, '')
                        
                        if not value:
                            italian_key = next((k for k, v in field_mapping.items() if v == col), None)
                            if italian_key:
                                value = tc_data.get(italian_key, '')
                        
                        row[col] = value
                first = False
            else:
                for col in columns:
                    if col not in ['Step', 'Step Description', 'Step Expected Result']:
                        row[col] = ''

            row['Step'] = step.get('Step', '')
            row['Step Description'] = step.get('Step Description', '')
            row['Step Expected Result'] = step.get('Expected Result', '')
            
            rows.append(row)

    df = pd.DataFrame(rows, columns=columns)


    excel_path = os.path.join(os.path.dirname(__file__), "..", "outputs", "testbook_controllosintattico_feedbackAI.xlsx")
    os.makedirs(os.path.dirname(excel_path), exist_ok=True)

    df.to_excel(excel_path, index=False)

    wb = load_workbook(excel_path)
    ws = wb.active

    for row in ws.iter_rows(min_row=2):  
        for cell in row:
            if cell.value and isinstance(cell.value, str) and "[[RED]]" in cell.value:
                apply_red_text(cell)

    wb.save(excel_path)
    print(f"Excel salvato con testi rossi: {excel_path}")




def convert_json_to_excel(json_data, output_path):
    """
    Converte il JSON con test_cases in un file Excel,
    con gli step su righe separate (stile Polarion)
    """
    
    # Definisci colonne finali
    columns = [
        'Title', 'ID', '#', 'Test Group', 'Channel', 'Device', 
        'Priority', 'Test Stage', 'Reference System', 
        'Preconditions', 'Execution Mode', 'Functionality', 
        'Test Type', 'No Regression Test', 'Automation',
        'Dataset', 'Expected Result', 
        'Step', 'Step Description', 'Step Expected Result',
        'Country', 'Project', 'Author', 'Assignee(s)', 'Type', 
        'Partial Coverage Description', '_polarion',
        'Analysis', 'Coverage', 'Dev Complexity', 'Execution Time', 
        'Volatility', 'Developed', 'Note', 'Team Ownership', 
        'Team Ownership Note', 'Requires Script Maintenance'
    ]

    rows = []
    
    # Estrai l'array test_cases dal JSON
    test_cases = json_data.get('test_cases', [])
    counter=1
    
    for tc_data in test_cases:
        steps = tc_data.get('Steps', [])
        
        # Se non ci sono step, crea una riga vuota
        if not steps:
            steps = [{}]
        
        first = True
        for step in steps:
            row = {}

            # Prima riga: compila tutti i campi del test case
            if first:
                for col in columns:
                    if col not in ['Step', 'Step Description', 'Step Expected Result']:
                        row[col] = tc_data.get(col, '')
                row['#'] = counter
                counter += 1
                first = False
            else:
                # Righe successive: lascia vuoti i campi del test case
                for col in columns:
                    if col not in ['Step', 'Step Description', 'Step Expected Result']:
                        row[col] = ''

            # Compila i campi dello step
            row['Step'] = step.get('Step', '')
            row['Step Description'] = step.get('Step Description', '')
            row['Step Expected Result'] = step.get('Expected Result', '')
            
            rows.append(row)

    # Crea DataFrame
    df = pd.DataFrame(rows, columns=columns)

    # Salva Excel
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    df.to_excel(output_path, index=False)
    
    # Opzionale: formatta il file Excel
    wb = load_workbook(output_path)
    ws = wb.active
    
    # Auto-adatta larghezza colonne (opzionale)
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(output_path)
    print(f"Excel salvato: {output_path}")




def from_en_to_it(testcase):



    ### NOTA EXPECTED RESULT NON VA BENE
    
    mapping = {
    "Title": "Title",  
    "ID": "ID",  
    "#": "#",  
    "Test Group": "Test Group",
    "Channel": "Canale",
    "Device": "Dispositivo",
    "Priority": "Priority",
    "Test Stage": "Test Stage",
    "Reference System": "Sistema di riferimento",
    "Preconditions": "Precondizioni",
    "Execution Mode": "Modalità Operativa",
    "Functionality": "Funzionalità",
    "Test Type": "Tipologia Test",
    "No Regression Test": "Test di no regression",
    "Automation": "Automation",
    "Dataset": "Dataset",
    "Expected Result": "Risultato Atteso",
    "Step": "Step",  
    "Step Description": "Step Description",  
    "Expected Result": "Expected Result",  
    "Country": "Country",
    "Project": "Project",
    "Author": "Author",
    "Assignee(s)": "Assignee(s)",
    "Type": "Type",
    "Partial Coverage Description": "Description Partial Coverage",
    "_polarion": "_polarion",
    "Analysis": "Analysis",
    "Coverage": "Coverage",
    "Dev Complexity": "Dev Complexity",
    "Execution Time": "Execution Time",
    "Volatility": "Volatility",
    "Developed": "Developed",
    "Note": "Note",
    "Team Ownership": "Team Ownership",
    "Team Ownership Note": "Team Ownership Note",
    "Requires Script Maintenance": "Requires Script Maintenance"
}


# import docx2pdf
from pdf2image import convert_from_path
import os

def docx_to_image(docx_path, output_folder=None, dpi=200):
    """
    Converte un file DOCX in immagini (una per pagina)
    
    Args:
        docx_path: percorso del file DOCX
        output_folder: cartella di output (default: stessa del DOCX)
        dpi: risoluzione delle immagini (default: 200)
    
    Returns:
        lista dei percorsi delle immagini create
    """
    
    # Imposta cartella di output
    if output_folder is None:
        output_folder = os.path.dirname(docx_path)
    
    os.makedirs(output_folder, exist_ok=True)
    
    # Nome base del file
    base_name = os.path.splitext(os.path.basename(docx_path))[0]
    
    # Step 1: Converti DOCX in PDF temporaneo
    pdf_path = os.path.join(output_folder, f"{base_name}_temp.pdf")
    print(f"Conversione {docx_path} in PDF...")
    docx2pdf.convert(docx_path, pdf_path)
    
    # Step 2: Converti PDF in immagini
    print(f"Conversione PDF in immagini...")
    images = convert_from_path(pdf_path, dpi=dpi)
    
    # Step 3: Salva le immagini
    image_paths = []
    for i, img in enumerate(images, 1):
        img_path = os.path.join(output_folder, f"{base_name}_pagina_{i}.png")
        img.save(img_path, 'PNG')
        image_paths.append(img_path)
        print(f"Salvata: {img_path}")
    
    # Rimuovi PDF temporaneo
    os.remove(pdf_path)
    
    print(f"\nConversione completata! Create {len(image_paths)} immagini.")
    return image_paths





def estrai_immagini_da_docx(docx_path, output_folder="image"):
    """
    Estrae tutte le immagini da un file DOCX
    
    Args:
        docx_path: percorso del file DOCX
        output_folder: cartella di output (default: docx_images)
    
    Returns:
        lista dei percorsi delle immagini estratte
    """
    
    # Imposta cartella di output
    if output_folder is None:
        output_folder = "image"
    
    os.makedirs(output_folder, exist_ok=True)
    
    # Apri il documento
    doc = Document(docx_path)
    
    # Nome base del file
    base_name = os.path.splitext(os.path.basename(docx_path))[0]
    
    image_paths = []
    image_count = 0
    
    # Estrai le immagini dalle relazioni del documento
    for rel in doc.part.rels.values():
        if "image" in rel.target_ref:
            image_count += 1
            
            # Ottieni i dati dell'immagine
            image_data = rel.target_part.blob
            
            # Determina l'estensione del file
            content_type = rel.target_part.content_type
            ext = content_type.split('/')[-1]
            if ext == 'jpeg':
                ext = 'jpg'
            
            # Salva l'immagine
            img_filename = f"{base_name}_immagine_{image_count}.{ext}"
            img_path = os.path.join(output_folder, img_filename)
            
            with open(img_path, 'wb') as img_file:
                img_file.write(image_data)
            
            image_paths.append(img_path)
            print(f"Estratta: {img_filename}")
    
    if image_count == 0:
        print("Nessuna immagine trovata nel documento. provo a convertire il docx in pdf e poi estrarre le immagini dal pdf")

        image_paths = docx_to_image(
            docx_path=docx_path,
            output_folder="image",  # cartella di output
            dpi=300  # qualità alta
        )

    else:
        print(f"\nTotale immagini estratte: {image_count}")
    
    return image_paths


def estrai_immagini_da_pdf(pdf_path, output_folder="image", dpi=200):

    base_name = os.path.splitext(os.path.basename(pdf_path))[0]
    images = convert_from_path(pdf_path, dpi=dpi)
    
    # Step 3: Salva le immagini
    image_paths = []
    for i, img in enumerate(images, 1):
        img_path = os.path.join(output_folder, f"{base_name}_paginacreata_{i}.png")
        img.save(img_path, 'PNG')
        image_paths.append(img_path)
        print(f"Salvata: {img_path}")
    
    # Rimuovi PDF temporaneo
    os.remove(pdf_path)
    
    print(f"\nConversione completata! Create {len(image_paths)} immagini.")
    return image_paths